# -*- coding: utf-8 -*-
"""
@author: flySky
"""
import os
import time
import hmac
import hashlib
import base64
import urllib.parse
import requests
import json
import threading
import queue
import urllib3
from openpyxl import Workbook
import openpyxl

# 移除报错
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
requests.packages.urllib3.disable_warnings()
requests.packages.urllib3.util.ssl_.DEFAULT_CIPHERS += 'HIGH:!DH:!aNULL'


class LiveTest(threading.Thread):
    def __init__(self, queue1):
        threading.Thread.__init__(self)
        self._queue = queue1

    def run(self):
        while True:
            if self._queue.empty():
                break
            # 实际工作的代码区域
            headers = {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36',
                'Connection': 'close'}
            url = self._queue.get(timeout=0.5)
            if url[:4] == 'http':
                try:
                    response = requests.get(url, headers=headers, timeout=60, verify=False).status_code
                    response_time = requests.get(url, headers=headers, timeout=60,
                                                 verify=False).elapsed.microseconds / 1000
                    # print(response)
                    # print(response_time)
                    if response == 400:
                        try:
                            url1 = 'https://' + url
                            response = requests.get(url1, headers=headers, timeout=60, verify=False).status_code
                            response_time = requests.get(url1, headers=headers, timeout=60,
                                                         verify=False).elapsed.microseconds / 1000
                            self.OutPut(url1, str(response), str(response_time))
                            continue
                        except Exception as e:

                            self.OutPut(url, f'Error2: {str(e)}', response_time)
                            continue
                    else:
                        self.OutPut(url, str(response), str(response_time))
                    continue
                except Exception as e:
                    self.OutPut(url, f'timeout', "10000")
                    continue
            else:
                try:
                    url1 = 'http://' + url
                    response = requests.get(url1, headers=headers, timeout=60, verify=False).status_code
                    response_time = requests.get(url, headers=headers, timeout=60,
                                                 verify=False).elapsed.microseconds / 1000
                    if response == 400:
                        try:
                            url1 = 'https://' + url
                            response = requests.get(url1, headers=headers, timeout=60, verify=False).status_code
                            response_time = requests.get(url1, headers=headers, timeout=60,
                                                         verify=False).elapsed.microseconds / 1000
                            self.OutPut(url1, str(response), str(response_time))
                            continue
                        except Exception as e:
                            self.OutPut(url, f'Error2: {str(e)}')
                            continue
                    else:
                        self.OutPut(url1, str(response), str(response_time))
                    continue
                except requests.exceptions.ConnectionError:
                    try:
                        url2 = 'https://' + url
                        response = requests.get(url2, headers=headers, timeout=60, verify=False).status_code
                        response_time = requests.get(url1, headers=headers, timeout=60,
                                                     verify=False).elapsed.microseconds / 1000
                        self.OutPut(url1, str(response), str(response_time))
                        continue
                    except Exception as e:
                        self.OutPut(url, f'Error2: {str(e)}')
                        continue
                except Exception as e:
                    self.OutPut(url, f'Error3: {str(e)}')
                    continue

    def OutPut(self, url, stat, time):
        lock.acquire()
        # print("url:\t" + url, "\t\t\tcode:\t" + stat, "\ttime:\t"+time + "ms")
        print("url:%-50s\tcode:%-10s\ttime:%-10s" % (url, stat, time))
        Url_stat.append([url, stat, time])
        lock.release()


def dingDing(send_data):
    timestamp = str(round(time.time() * 1000))
    # secret = 'SEC71d6f32ae6c993c28ef83cc56f3dce89da92c2b3446c9ee12b97254d93e45755'
    secret = dingding_robot_secret
    secret_enc = secret.encode('utf-8')
    string_to_sign = '{}\n{}'.format(timestamp, secret)
    string_to_sign_enc = string_to_sign.encode('utf-8')
    hmac_code = hmac.new(secret_enc, string_to_sign_enc, digestmod=hashlib.sha256).digest()
    sign = urllib.parse.quote_plus(base64.b64encode(hmac_code))
    access_token = 'a29f4414fbc19ef15464183442995cbc57bbb23c6cd021af8371796adc538278'
    url = 'https://oapi.dingtalk.com/robot/send?access_token={}&timestamp={}&sign={}'.format(access_token, timestamp,
                                                                                             sign)
    # print(url)
    str_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
    data = {"msgtype": "text",
            "text": {
                "title": "web站点延迟测试报告",
                "content": "业务报警：\n告警时间：" + str_time + "\n" + send_data
                # "content": "业务报警:\n时间:{}\nweb站点{}\n当前站点访问延迟:{}\n".format(str_time, monitor_url, response_time)
            },
            "at": {"atMobiles": phone_num_list}
            # "at": {"atMobiles": ["18236593202"]}
            }
    headers = {'Content-Type': 'application/json'}
    message = requests.post(url, json.dumps(data), headers=headers).json()
    print(message)
    if message['errmsg'] == 'ok':
        print('钉钉消息推送成功')
    else:
        print('钉钉消息推送失败')


def write_status_to_file(file_name_stat, Url_stat):
    curent_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
    # file_name_stat = 'web_state_recording_' + time.strftime('%Y-%m-%d', time.localtime()) + ".xlsx"
    # print(Url_stat)
    for i in Url_stat:
        i.insert(0, curent_time)

    if os.path.exists(file_name_stat):
        workbook = openpyxl.load_workbook(file_name_stat)
        # sheet = workbook.get_sheet_by_name("UrlStat")
        sheet = workbook["UrlStat"]
        for web_list in Url_stat:
            sheet.append(web_list)
        workbook.save(file_name_stat)
    else:
        workbook = Workbook()
        workbook.create_sheet("UrlStat", 0)
        sheet = workbook.active
        sheet["A1"] = "检测时间"
        sheet["B1"] = "webUrl"
        sheet["C1"] = "状态码"
        sheet["D1"] = "url响应时间"
        for web_list in Url_stat:
            sheet.append(web_list)
        workbook.save(file_name_stat)


def main(filepath):
    global lock, worksheet, Url_stat
    Url_stat = []  # 用来存放访问结果，最后存入xls
    xujiancedeURL = open(filepath, 'r', encoding='utf-8')  # 读取所要检测的url列表。就是提供给核心代码区域的参数
    thread_count = 1000  # 线程数
    threads = []
    queue1 = queue.Queue()
    lock = threading.Lock()
    for line in xujiancedeURL:
        line = line.strip()  # 读取每一行
        queue1.put(line)
    for i in range(thread_count):
        threads.append(LiveTest(queue1))
    for t in threads:
        t.start()
    for t in threads:
        t.join()
    xujiancedeURL.close()
    # print(Url_stat)
    file_name_stat = 'web_state_recording_' + time.strftime('%Y-%m-%d', time.localtime()) + ".xlsx"
    write_status_to_file(file_name_stat, Url_stat)

    # 延迟超过阈值发送dingding告警 修改300毫秒的值即可
    exceed_threshold_value_list = [n for n in Url_stat if float(n[3]) > send_notice_time_threshold_value]
    send_tmp_data = ["weburl \t" + x[1].ljust(60, ' ') + "响应时间\t" + x[3].ljust(10, ' ') + "\n" for x in
                     exceed_threshold_value_list]
    txt_info = "".join(send_tmp_data)
    dingDing(txt_info)


if __name__ == '__main__':
    # 设置响应时间告警阈值
    send_notice_time_threshold_value = 300
    # 钉钉机器人 secret
    dingding_robot_secret = "SEC71d6f32ae6c993c28ef83cc56f3dce89da92c2b3446c9ee12b97254d93e45755"
    # 通知人电话列表
    phone_num_list = ["18236593202"]
    main(r"D:\url.txt")
